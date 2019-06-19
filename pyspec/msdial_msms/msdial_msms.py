"""MS-Dial MSMS feature scraper

This script finds and stores information on all features that contain MSMS in the raw export sheet from MS-Dial.
It automatically opens the single .xlsx file in the same directory as the script and stores all data
into a tuple called feature.  Each feature has the following data which can be accessed as follows:

feature[num].retention_time - feature retention time
feature[num].mz - feature mass to charge ratio
feature[num].msms - tuple of type msms spectra in the format double:int (mz:peak height)

num can be 0 and len(feature).

This file can also be imported as a module and contains the following functions:
    *getExcelSheets - Opens all excel sheets in directory of .py script and put them into a list.
    *getFileName - Returns file name from list of excel file names.
    *openWorkBook - Opens excel workbook of excel sheet chosen from excel sheet list.
    *makeSheet - Returns first sheet of excel workbook passed in.
    *msmsFound - Returns True if current row and column 8 of excel sheet value is 'True'.
    *retentionTimeColumn - Returns True if row 4 and column 8 of excel sheet value is 'Average Rt(min)'.
    *mzColumn - Returns True if row 4 and column 3 of excel sheet value is 'Average Mz'.
    *msmsColumn - Returns True if row 4 and column 3 of excel sheet value is 'MS/MS spectrum'.
    *collectFeatures - For all rows with MSMS data, stores feature objects in feature tuple.
    *main - The main function of the script.

Author: Bryan Roberts
"""

import openpyxl
import os

def getExcelSheets():
    """Opens all excel sheets in directory of .py script and put them into a list
            
    Returns
    -------
    list
        A list of all excel files in current directory.  Item names are full directory path
        plus file name.
    """
    
    excelSheets = []
    for file in os.listdir():
        if file[-5:] == '.xlsx':
            if file[0] != '~':
                excelSheets.append(os.path.join(os.getcwd(), file))
    return excelSheets

def getFileName(excelSheets, index):
    """Returns file name from list of excel file names
            
    Parameters
    ----------
    excelSheets : list
        A list of all excel files in current directory.
    index : int
        index of excel sheets list to extract file name from
        
    Returns
    -------
    str
        string file name minus full directory path
    """
    
    split = excelSheets[index].split(os.path.sep)
    return split[-1]

def openWorkBook(excelSheets, index):
    """Opens excel workbook of excel sheet chosen from excel sheet list
            
    Parameters
    ----------
    excelSheets : list
        A list of all excel files in current directory.
    index : int
        index of excel sheets list to extract file name from
        
    Returns
    -------
    wb
        Workbook of chosen excel sheet
    """
    
    wb = openpyxl.load_workbook(excelSheets[index])
    return wb


def makeSheet(wb):
    """Returns first sheet of excel workbook passed in.
            
    Parameters
    ----------
    wb : workbook 
        Workbook opened using openpyxl
    
    Returns
    -------
    Workbook sheet
        returns first sheet in excel file opened  
    """
    
    sheets = wb.sheetnames
    sheet = wb[sheets[0]]
    return sheet

def msmsFound(sheet, currentRow, currentColumn = 8):
    """Returns True if current row and column 8 of excel sheet value is 'True'
           
    Parameters
    ----------
    sheet : excel workbook sheet
        The first sheet in excel workbook opened
    currentRow : int
        The current row of the feature that data is being taken from
    currentColumn : int
        The column in the MS-Dial export sheet that contains msms true of false.
        Default value of 8.
            
    Returns
    -------
    bool
        True if cell value == 'True', otherwise False
    """
    
    return str(sheet.cell(row = currentRow, column = currentColumn).value) == 'True'

def retentionTimeColumn(sheet):
    """Returns True if row 4 and column 8 of excel sheet value is 'Average Rt(min)'
            
    Parameters
    ----------
    sheet : excel workbook sheet
        The first sheet in excel workbook opened
            
    Returns
    -------
    bool
        True if cell value == 'Average Rt(min)', otherwise False
    """
    
    return sheet.cell(row = 4, column = 2).value == 'Average Rt(min)'

def mzColumn(sheet):
    """Returns True if row 4 and column 3 of excel sheet value is 'Average Mz'
            
    Parameters
    ----------
    sheet : excel workbook sheet
        The first sheet in excel workbook opened
            
    Returns
    -------
    bool
        True if cell value == 'Average Mz', otherwise False
    """
    
    return sheet.cell(row = 4, column = 3).value == 'Average Mz'

def msmsColumn(sheet):
    """Returns True if row 4 and column 3 of excel sheet value is 'MS/MS spectrum'
            
    Parameters
    ----------
    sheet : excel workbook sheet
        The first sheet in excel workbook opened
            
    Returns
    -------
    bool
        True if cell value == 'MS/MS spectrum', otherwise False
    """
    
    return sheet.cell(row = 4, column = 22).value == 'MS/MS spectrum'
    
def test_retentionTimeColumn(sheet):
    """py_test to make sure retention time column is correct in excel file opened.
            
    Parameters
    ----------
    sheet : excel workbook sheet
        The first sheet in excel workbook opened
    """
    
    assert retentionTimeColumn(sheet) == True
    
def test_mzColumn(sheet):
    """py_test to make sure mass to charge column is correct in excel file opened.
            
    Parameters
    ----------
    sheet : excel workbook sheet
        The first sheet in excel workbook opened
    """
    
    assert mzColumn(sheet) == True
    
def test_msmsColumn(sheet):
    """py_test to make sure msms column is correct in excel file opened.
            
    Parameters
    ----------
    sheet : excel workbook sheet
        The first sheet in excel workbook opened
    """
    
    assert msmsColumn(sheet) == True

def collectFeatures(sheet, filename):
    """For all rows with MSMS data, stores feature objects in feature tuple
            
    Parameters
    ----------
    sheet : excel workbook sheet
        The first sheet in excel workbook opened
    filename : str
        The file name of the excel document being worked with
            
    Returns
    -------
    tuple
        Tuple of class Feature objects
    """
    
    features = ()#initiate empty tuple
    #count = 0
    current_row = 5
    
    #py_tests to check excel sheet is correct format
    test_retentionTimeColumn(sheet)
    test_mzColumn(sheet)
    test_msmsColumn(sheet)
    
    #check all rows for msms data and store data if msms data is present
    while (current_row <= sheet.max_row):
        if msmsFound(sheet, current_row):
            feature = Feature(sheet, current_row)
            features = features + (feature,)
            #count = count + 1
        
    print(str(len(features)) + ' features with MSMS found in ' + filename)
    return features
        
class Feature():
    """A class used to represent individual feature data from raw MS-Dial exported sheet
    
    Attributes
    ----------
    retention_time : double
        'Average Rt(min)' column in export sheet (column = 2)
    mz : double
        'Average Mz' column in export sheet (column = 3)
    msms : tuples of type spectra
        'MS/MS spectrum' column in export sheet (column = 22)
        format - mz:height
    
    Methods
    -------
    setRetentionTime(self, sheet, currentRow, currentColumn = 2)
        Sets retention_time to value found in currentrow and column 2
    setMZ(self, sheet, currentRow, currentColumn = 3)
        Sets mz to value found in current row and column 3
    setMSMS(self, sheet, currentRow, currentColumn = 22)
        Sets msms to tuple of values found in current row and column 22
    """
    
    def __init__(self, sheet, currentRow):
        """
        Parameters
        ----------
        sheet : excel workbook sheet
            The first sheet in excel workbook opened
        currentRow : int
            The current row of the feature that data is being taken from
        """
        
        self.retention_time = self.setRetentionTime(sheet, currentRow)
        self.mz = self.setMZ(sheet, currentRow)
        self.msms = self.setMSMS(sheet, currentRow)
        
    def setRetentionTime(self, sheet, currentRow, currentColumn = 2):
        """Sets retention_time to value found in currentrow and column 2
        
        Parameters
        ----------
        sheet : excel workbook sheet
            The first sheet in excel workbook opened
        currentRow : int
            The current row of the feature that data is being taken from
        currentColumn : int
            The column in the MS-Dial export sheet that contains retention time information.
            Default value of 2.
            
        Returns
        -------
        double
            The retention time for the feature on the current row
        """
        
        return sheet.cell(row = currentRow, column = currentColumn).value
    
    def setMZ(self, sheet, currentRow, currentColumn = 3):
        """Sets mz (mass to charge) to value found in currentrow and column 2
        
        Parameters
        ----------
        sheet : excel workbook sheet
            The first sheet in excel workbook opened
        currentRow : int
            The current row of the feature that data is being taken from
        currentColumn : int
            The column in the MS-Dial export sheet that contains mass to charge information.
            Default value of 3.
            
        Returns
        -------
        double
            The mass to charge for the feature on the current row
        """
        
        return sheet.cell(row = currentRow, column = 3).value
    
    def setMSMS(self, sheet, currentRow, currentColumn = 22):
        """Sets msms (MS2 spectra) to value found in currentrow and column 22 into tuple
        
        Parameters
        ----------
        sheet : excel workbook sheet
            The first sheet in excel workbook opened
        currentRow : int
            The current row of the feature that data is being taken from
        currentColumn : int
            The column in the MS-Dial export sheet that contains msms information.
            Default value of 22.
            
        Returns
        -------
        tuple
            The msms spectra for the feature on the current row in a tuple
            format - mz:height
        """
        
        cell_content = sheet.cell(row = currentRow, column = currentColumn).value
        spectra = cell_content.split()
        return tuple(spectra)

def main(): 
    file = getExcelSheets()
    file_name = getFileName(file, 0)
    wb = openWorkBook(file, 0)
    sheet = makeSheet(wb)
    feature = collectFeatures(sheet, file_name)
    return feature

if __name__ == "__main__":
    feature = main() 
